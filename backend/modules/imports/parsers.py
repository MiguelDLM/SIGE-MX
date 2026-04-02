# backend/modules/imports/parsers.py
import csv
import io
from typing import Any

import openpyxl


def parse_file(content: bytes, filename: str) -> list[dict[str, Any]]:
    """Parse CSV or XLSX file and return list of row dicts (header = keys)."""
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")  # handle BOM
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        result.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    return result


def validate_student_row(row: dict[str, Any], row_num: int) -> list[dict]:
    """Return list of error dicts for a student row. Empty list = valid."""
    errors = []
    for field in ("nombre", "apellido_paterno", "matricula"):
        if not row.get(field, "").strip():
            errors.append({"row": row_num, "field": field, "message": f"Campo '{field}' es requerido"})
    return errors


def validate_teacher_row(row: dict[str, Any], row_num: int) -> list[dict]:
    """Return list of error dicts for a teacher row. Empty list = valid."""
    errors = []
    for field in ("nombre", "apellido_paterno", "numero_empleado"):
        if not row.get(field, "").strip():
            errors.append({"row": row_num, "field": field, "message": f"Campo '{field}' es requerido"})
    return errors


def build_student_template_xlsx() -> bytes:
    """Return bytes of an .xlsx template for student import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alumnos"
    headers = ["nombre", "apellido_paterno", "apellido_materno", "matricula",
               "municipio", "estado", "codigo_postal", "tipo_sangre"]
    ws.append(headers)
    ws.append(["Ana", "García", "López", "2024001", "Monterrey", "Nuevo León", "64000", "O+"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_teacher_template_xlsx() -> bytes:
    """Return bytes of an .xlsx template for teacher import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Docentes"
    headers = ["nombre", "apellido_paterno", "apellido_materno", "numero_empleado", "especialidad"]
    ws.append(headers)
    ws.append(["Carlos", "Mendoza", "Ruiz", "EMP001", "Matemáticas"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
